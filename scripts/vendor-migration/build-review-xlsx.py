"""협력업체 이관 검수표 생성 — 온전히 옮길 수 있는 업체만 추린다."""
import json
import re
import urllib.request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

W = 'C:/Users/dawoo/ERP 헤르메스/'
OUT = W + 'vendor-migration-work/협력업체_이관_검수표.xlsx'

read = json.load(open(W + 'vendor-migration-work/read.json', encoding='utf-8'))
fmap = json.load(open(W + 'dawoo-erp/scripts/vendor-migration/folder-map.json', encoding='utf-8'))['map']
scan = json.load(open(W + 'vendor-migration-work/scan.json', encoding='utf-8'))
files_by_folder = {v['folder']: v for v in scan['vendors']}

norm = lambda s: re.sub(r'주식회사|\(주\)|㈜|\s', '', s or '')

# 기존 ERP 거래처DB를 읽어온다. 판독이 못 채운 항목을 ERP가 이미 갖고 있으면
# 그 값으로 메운다 — 이관 대상을 억울하게 보류시키지 않기 위해서다.
env = open(W + 'dawoo-erp/.env.local', encoding='utf-8').read()
sb_url = re.search(r'^NEXT_PUBLIC_SUPABASE_URL="?([^"\r\n]+)"?', env, re.M).group(1)
sb_key = re.search(r'^SUPABASE_SERVICE_ROLE_KEY="?([^"\r\n]+)"?', env, re.M).group(1)
req = urllib.request.Request(
    sb_url + '/rest/v1/vendors?select=name,business_number,representative,address,bank_info'
             '&vendor_type=eq.%ED%98%91%EB%A0%A5%EC%97%85%EC%B2%B4',
    headers={'apikey': sb_key, 'Authorization': 'Bearer ' + sb_key})
erp_rows = json.load(urllib.request.urlopen(req))

def split_bank(raw):
    """vendorBank.ts와 같은 규칙 — '은행/계좌/예금주'."""
    parts = [p.strip() for p in (raw or '').split('/')]
    return {'bank_name': parts[0] if len(parts) > 0 else '',
            'account_number': parts[1] if len(parts) > 1 else '',
            'account_holder': parts[2] if len(parts) > 2 else ''}

ERP = {}
for v in erp_rows:
    ERP[norm(v['name'])] = {
        'company_name': v.get('name') or '', 'business_number': v.get('business_number') or '',
        'representative': v.get('representative') or '', 'address': v.get('address') or '',
        **split_bank(v.get('bank_info')),
    }
EXISTING = [v['name'] for v in erp_rows]

CORE = ['company_name', 'business_number', 'representative',
        'address', 'bank_name', 'account_number', 'account_holder']

ready, hold = [], []
for folder, r in read.items():
    res = dict(r.get('result') or {})
    m = fmap[folder]
    erp = ERP.get(norm(res.get('company_name', ''))) or ERP.get(norm(m['name']))
    filled_from_erp = []
    if erp:
        for f in CORE:
            if not res.get(f) and erp.get(f):
                res[f] = erp[f]
                filled_from_erp.append(f)
    if all(res.get(f) for f in CORE):
        ready.append((folder, r, res, m, filled_from_erp))
    else:
        missing = [f for f in CORE if not res.get(f)]
        hold.append((folder, r, res, m, missing))

ready.sort(key=lambda x: x[2]['company_name'])
hold.sort(key=lambda x: x[0])
LABEL = {'company_name': '상호', 'business_number': '사업자번호', 'representative': '대표자',
         'address': '주소', 'bank_name': '은행', 'account_number': '계좌번호',
         'account_holder': '예금주'}

# --- 스타일 ---
ARIAL = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(name=ARIAL, bold=True, color='FFFFFF', size=10)
REF_FILL = PatternFill('solid', fgColor='F2F2F2')   # 참고용 — 고치지 않음
CHK_FILL = PatternFill('solid', fgColor='FFFF00')   # 확인란
DUP_FILL = PatternFill('solid', fgColor='FCE4D6')   # 기존 ERP와 중복
BODY = Font(name=ARIAL, size=10)
THIN = Side(style='thin', color='BFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ========== 시트 1: 안내 ==========
ws = wb.active
ws.title = '안내'
guide = [
    ('협력업체 이관 검수표', 14, True),
    ('', 10, False),
    ('무엇인가', 11, True),
    ('파일서버 "관리부 > 01. 사무관리 > 5. 협력업체 관련서류" 폴더 102곳을 읽고,', 10, False),
    ('사업자등록증·통장사본을 판독해 정리한 표입니다.', 10, False),
    ('이 중 상호·사업자번호·대표자·주소·은행·계좌·예금주 7개가 모두 채워진 곳만 [이관대상] 시트에 담았습니다.', 10, False),
    ('', 10, False),
    ('무엇을 하면 되나', 11, True),
    ('1. [이관대상] 시트를 위에서부터 훑으면서 값이 맞는지 봅니다.', 10, False),
    ('2. 틀린 값은 그 칸에서 바로 고치면 됩니다. 고친 값이 최종값입니다.', 10, False),
    ('3. 다 본 행은 맨 오른쪽 노란 [확인] 칸에 O 를 적습니다.', 10, False),
    ('4. ERP에 넣지 말아야 할 업체는 [확인] 칸에 X 를 적습니다.', 10, False),
    ('5. 저장해서 돌려주시면 그대로 ERP에 등록합니다.', 10, False),
    ('', 10, False),
    ('색깔 규칙', 11, True),
    ('회색 칸  — 참고용입니다. 고치지 마세요. (원본폴더명, 판독근거)', 10, False),
    ('흰 칸    — 고쳐도 되는 칸입니다. (업체명 ~ 비고)', 10, False),
    ('노란 칸  — 확인 표시를 적는 칸입니다.', 10, False),
    ('주황 행  — 이미 ERP 거래처DB에 있는 업체입니다. 기존 값은 덮어쓰지 않고 빈칸만 채웁니다.', 10, False),
    ('', 10, False),
    ('꼭 대조해 주실 것', 11, True),
    ('사업자번호와 계좌번호는 숫자 하나만 틀려도 송금 사고가 납니다.', 10, False),
    ('[판독근거] 칸의 파일 경로를 열어 원본과 대조해 주세요.', 10, False),
    ('', 10, False),
    ('보류 시트', 11, True),
    ('서류가 없거나 일부만 읽혀서 이번에 넣지 않은 업체 목록입니다.', 10, False),
    ('서류를 구해오시면 그때 추가로 넣을 수 있습니다.', 10, False),
]
for i, (text, size, bold) in enumerate(guide, start=1):
    c = ws.cell(row=i, column=1, value=text)
    c.font = Font(name=ARIAL, size=size, bold=bold)
ws.column_dimensions['A'].width = 110

# 예시 행 — 어떤 형식으로 적으면 되는지
r0 = len(guide) + 2
ws.cell(row=r0, column=1, value='적는 형식 예시').font = Font(name=ARIAL, size=11, bold=True)
ex = [('업체명', '(주)삼진공조이엔지'), ('공종', '냉난방기, 설비'),
      ('사업자번호', '310-86-02312'), ('대표자', '김옥경'),
      ('은행', '우리은행'), ('계좌번호', '1005-804-125152'),
      ('예금주', '(주)삼진공조이엔지'), ('전화', '031-719-8703'), ('확인', 'O')]
for j, (k, v) in enumerate(ex):
    ws.cell(row=r0 + 1 + j, column=1, value=f'  {k}  →  {v}').font = Font(name=ARIAL, size=10)

# ========== 시트 2: 이관대상 ==========
COLS = [
    ('번호', 6, 'ref'), ('원본폴더명', 34, 'ref'), ('업체명', 24, 'edit'),
    ('공종', 22, 'edit'), ('사업자번호', 15, 'edit'), ('대표자', 12, 'edit'),
    ('주소', 46, 'edit'), ('은행', 14, 'edit'), ('계좌번호', 20, 'edit'),
    ('예금주', 20, 'edit'), ('전화', 15, 'edit'), ('이메일', 24, 'edit'),
    ('비고', 40, 'edit'), ('첨부파일', 9, 'ref'), ('판독근거', 60, 'ref'), ('확인', 8, 'chk'),
]

ws2 = wb.create_sheet(f'이관대상({len(ready)})')
for j, (name, width, _) in enumerate(COLS, start=1):
    c = ws2.cell(row=1, column=j, value=name)
    c.fill, c.font = HDR_FILL, HDR_FONT
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.column_dimensions[get_column_letter(j)].width = width
ws2.row_dimensions[1].height = 22

for i, (folder, r, res, m, from_erp) in enumerate(ready, start=1):
    email = res.get('email', '')
    notes = []
    if from_erp:
        notes.append('※ ' + ', '.join(LABEL[f] for f in from_erp) + ' 은(는) 서류에서 못 읽어 기존 ERP 값을 그대로 씀 — 대조 필요')
    if m.get('note'):
        notes.append(m['note'])
    # 국세청 전용주소는 연락용 이메일이 아니다 — 칸에서 빼고 비고로 옮긴다.
    if 'hometax' in email.lower():
        notes.append(f'세금계산서 전용메일 {email} (연락용 아님)')
        email = ''
    if res.get('notes'):
        notes.append(res['notes'])

    is_dup = any(norm(e) == norm(res['company_name']) or norm(e) == norm(m['name']) for e in EXISTING)
    if is_dup:
        notes.insert(0, '★ 기존 ERP 거래처DB에 이미 있음 — 빈칸만 채움')

    nfiles = len([f for f in files_by_folder[folder]['files'] if not f.get('skip')])
    row = [
        i, folder, res['company_name'], ', '.join(m['categories']),
        res['business_number'], res['representative'], res['address'],
        res['bank_name'], res['account_number'], res['account_holder'],
        res.get('phone', ''), email, ' / '.join(notes), nfiles,
        ' | '.join(p.split('\\')[-1] for p in r.get('sources', [])), '',
    ]
    for j, val in enumerate(row, start=1):
        c = ws2.cell(row=i + 1, column=j, value=val)
        c.font, c.border = BODY, BOX
        c.alignment = Alignment(vertical='top', wrap_text=(j in (7, 13, 15)))
        kind = COLS[j - 1][2]
        if kind == 'ref':
            c.fill = REF_FILL
        elif kind == 'chk':
            c.fill = CHK_FILL
        if is_dup and kind == 'edit':
            c.fill = DUP_FILL

ws2.freeze_panes = 'C2'
ws2.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}{len(ready) + 1}'

# ========== 시트 3: 보류 ==========
HCOLS = [('원본폴더명', 40), ('폴더명 기준 업체명', 22), ('공종', 20),
         ('사유', 14), ('빠진 항목', 34), ('읽힌 항목', 34), ('폴더 내 파일수', 12)]
ws3 = wb.create_sheet(f'보류({len(hold)})')
for j, (name, width) in enumerate(HCOLS, start=1):
    c = ws3.cell(row=1, column=j, value=name)
    c.fill, c.font = HDR_FILL, HDR_FONT
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.column_dimensions[get_column_letter(j)].width = width
ws3.row_dimensions[1].height = 22

for i, (folder, r, res, m, missing) in enumerate(hold, start=1):
    got = [LABEL[f] for f in CORE if res.get(f)]
    reason = '서류없음' if r.get('status') != 'ok' else '일부판독불가'
    nfiles = len([f for f in files_by_folder[folder]['files'] if not f.get('skip')])
    row = [folder, m['name'], ', '.join(m['categories']), reason,
           ', '.join(LABEL[f] for f in missing), ', '.join(got) or '(없음)', nfiles]
    for j, val in enumerate(row, start=1):
        c = ws3.cell(row=i + 1, column=j, value=val)
        c.font, c.border = BODY, BOX
        c.alignment = Alignment(vertical='top', wrap_text=(j in (5, 6)))
ws3.freeze_panes = 'A2'
ws3.auto_filter.ref = f'A1:{get_column_letter(len(HCOLS))}{len(hold) + 1}'

wb.save(OUT)
print(f'이관대상 {len(ready)}건 / 보류 {len(hold)}건')
dups = [res['company_name'] for f, r, res, m, _ in ready
        if any(norm(e) == norm(res['company_name']) or norm(e) == norm(m['name']) for e in EXISTING)]
print(f'기존 ERP와 중복: {len(dups)}건 — {", ".join(dups)}')
merged = [(res['company_name'], [LABEL[x] for x in fe]) for f, r, res, m, fe in ready if fe]
print(f'ERP 값으로 메워 이관대상이 된 곳: {len(merged)}건')
for nm, fs in merged:
    print(f'   {nm} ← {", ".join(fs)}')
print(OUT)
