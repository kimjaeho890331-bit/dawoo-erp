"""
④ 적재 — 검수 끝난 엑셀을 그대로 ERP에 넣는다.

설계: docs/superpowers/specs/2026-07-31-vendor-nas-migration-design.md
- 값의 최종 출처는 검수표 엑셀이다. read.json이 아니다.
- 기존 거래처는 빈칸만 채우고 기존 값은 덮어쓰지 않는다.
- NAS는 읽기 전용. 원본을 옮기거나 지우지 않는다.

실행:
  python scripts/vendor-migration/load.py            # 미리보기만 (쓰기 없음)
  python scripts/vendor-migration/load.py --apply    # 실제 적재
"""
import json
import mimetypes
import re
import sys
import urllib.error
import urllib.request
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

W = Path('C:/Users/dawoo/ERP 헤르메스')
WORK = W / 'vendor-migration-work'
XLSX = WORK / '협력업체_이관_검수표.xlsx'
SCAN = WORK / 'scan.json'
APPLY = '--apply' in sys.argv

env = (W / 'dawoo-erp/.env.local').read_text(encoding='utf-8')
URL = re.search(r'^NEXT_PUBLIC_SUPABASE_URL="?([^"\r\n]+)"?', env, re.M).group(1)
KEY = re.search(r'^SUPABASE_SERVICE_ROLE_KEY="?([^"\r\n]+)"?', env, re.M).group(1)
H = {'apikey': KEY, 'Authorization': 'Bearer ' + KEY}


def api(method, path, body=None, extra=None, raw=None, ctype=None):
    headers = dict(H)
    if extra:
        headers.update(extra)
    if raw is None:
        data = json.dumps(body).encode() if body is not None else None
        if data:
            headers['Content-Type'] = 'application/json'
    else:
        data = raw
        headers['Content-Type'] = ctype or 'application/octet-stream'
    req = urllib.request.Request(URL + path, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as r:
            text = r.read().decode('utf-8', 'replace')
            return json.loads(text) if text.strip().startswith(('{', '[')) else text
    except urllib.error.HTTPError as e:
        raise RuntimeError(f'{method} {path} → {e.code}: {e.read().decode("utf-8", "replace")[:400]}')


def safe_storage_path(p):
    """src/lib/utils/storagePath.ts와 같은 규칙."""
    return re.sub(r'[^a-zA-Z0-9._/-]', '_', p.replace('..', '').lstrip('/'))


norm = lambda s: re.sub(r'주식회사|\(주\)|㈜|\s', '', s or '')
squash = lambda s: re.sub(r'\s', '', s or '')

# --- 입력 ---
scan = json.loads(SCAN.read_text(encoding='utf-8'))
files_by_folder = {v['folder']: v['files'] for v in scan['vendors']}
FMAP = json.loads((W / 'dawoo-erp/scripts/vendor-migration/folder-map.json')
                  .read_text(encoding='utf-8'))['map']

wb = load_workbook(XLSX)
sheet = next(s for s in wb.sheetnames if s.startswith('이관대상'))
ws = wb[sheet]
head = [c.value for c in ws[1]]
col = {name: i for i, name in enumerate(head)}

rows, skipped = [], []
for r in range(2, ws.max_row + 1):
    vals = [ws.cell(row=r, column=c + 1).value for c in range(len(head))]
    rec = {name: (vals[i] or '') for name, i in col.items()}
    if str(rec['확인']).strip().upper() == 'X':
        skipped.append(rec['업체명'])
        continue
    rows.append(rec)

print(f'검수표 [{sheet}] {ws.max_row - 1}행 → 적재 대상 {len(rows)}건'
      + (f' / 제외(확인=X) {len(skipped)}건: {", ".join(skipped)}' if skipped else ''))

# --- 기존 거래처 ---
existing = api('GET', '/rest/v1/vendors?select=*&vendor_type=eq.%ED%98%91%EB%A0%A5%EC%97%85%EC%B2%B4')
by_name = {norm(v['name']): v for v in existing}
print(f'기존 협력업체 {len(existing)}건')

# --- 백업 ---
if APPLY:
    allv = api('GET', '/rest/v1/vendors?select=*')
    stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    bak = WORK / f'backup-vendors-{stamp}.json'
    bak.write_text(json.dumps(allv, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'백업: {bak}  ({len(allv)}건)')

# --- 공종칩 ---
# 운영 DB의 칩은 "금속 / 창호"처럼 슬래시 양옆에 공백이 있다. 내 매핑은 "금속/창호"다.
# 공백만 다른 같은 칩을 새로 만들면 같은 공종이 둘로 갈라지므로, 공백을 지워 대조하고
# 이미 있으면 DB 쪽 표기를 그대로 쓴다.
chips_have = [c['name'] for c in api('GET', '/rest/v1/vendor_categories?select=name')]
chip_alias = {squash(c): c for c in chips_have}


def chip(name):
    return chip_alias.get(squash(name), name)


for rec in rows:
    rec['공종'] = ', '.join(chip(c.strip()) for c in str(rec['공종']).split(',') if c.strip())

chips_want = {c.strip() for rec in rows for c in str(rec['공종']).split(',') if c.strip()}
chips_new = sorted(chips_want - set(chips_have))
reused = sorted({chip(c) for rec in rows for c in str(rec['공종']).split(',')
                 if c.strip() and squash(c.strip()) in chip_alias})
print(f'공종칩 신규 {len(chips_new)}개: {", ".join(chips_new)}')
print(f'기존 칩 재사용 {len(reused)}개: {", ".join(reused)}')
if APPLY and chips_new:
    api('POST', '/rest/v1/vendor_categories', [{'name': c} for c in chips_new],
        {'Prefer': 'return=minimal'})

# --- 서류 슬롯 ---
SLOTS = [('biz_license', 'biz_license_url'), ('bankbook', 'bankbook_url'),
         ('id_card', 'id_card_url'), ('safety_cert', 'safety_cert_url')]
UPLOADABLE = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf'}

# 남의 회사 서류가 섞여 들어간 폴더가 있다 (예: `※ 기원건설` 안에 산호이엔지 사업자등록증·통장).
# 파일명에 다른 업체 이름이 박혀 있으면 그 파일은 이 업체 서류로 쓰지 않는다.
OTHER_NAMES = {norm(FMAP[f]['name']) for f in FMAP}
OTHER_NAMES |= {norm(rec['업체명']) for rec in rows}
OTHER_NAMES = {n for n in OTHER_NAMES if len(n) >= 3}


def own_tokens(folder, name):
    t = {norm(name), norm(FMAP[folder]['name'])}
    return {x for x in t if len(x) >= 2}


def pick_for_slot(folder, name, tag):
    """이 업체 서류가 확실한 파일을 고른다. 애매하면 아예 고르지 않는다."""
    mine = own_tokens(folder, name)
    cands = [f for f in files_by_folder[folder]
             if not f.get('skip') and tag in f['tags'] and f['ext'] in UPLOADABLE]
    ok, rejected = [], []
    for f in cands:
        fn = norm(f['name'])
        is_mine = any(t in fn for t in mine)
        foreign = [o for o in OTHER_NAMES if o not in mine and o in fn]
        if foreign and not is_mine:
            rejected.append((f['name'], foreign[0]))
            continue
        ok.append((0 if is_mine else 1, f))  # 이름이 박힌 파일 우선
    ok.sort(key=lambda x: x[0])
    return (ok[0][1] if ok else None), rejected

log = {'ran_at': datetime.now().isoformat(), 'applied': APPLY,
       'inserted': [], 'updated': [], 'uploaded': [], 'errors': [],
       'conflicts': [], 'rejected_files': []}

for rec in rows:
    folder = rec['원본폴더명']
    name = str(rec['업체명']).strip()
    bank_info = '/'.join([str(rec['은행']).strip(), str(rec['계좌번호']).strip(),
                          str(rec['예금주']).strip()])
    note_parts = [f'NAS: {folder}']
    if rec['비고']:
        note_parts.append(str(rec['비고']))

    payload = {
        'name': name,
        'vendor_type': '협력업체',
        'category': str(rec['공종']).strip(),
        'business_number': str(rec['사업자번호']).strip(),
        'representative': str(rec['대표자']).strip(),
        # 화면(VendorsPage)은 담당자만 보여준다. 우선 대표자를 넣어 비어 보이지 않게 한다.
        'contact_person': str(rec['대표자']).strip(),
        'address': str(rec['주소']).strip(),
        'bank_name': str(rec['은행']).strip(),
        'account_number': str(rec['계좌번호']).strip(),
        'bank_info': bank_info,
        'phone': str(rec['전화']).strip(),
        'email': str(rec['이메일']).strip(),
        'note': ' / '.join(note_parts),
    }

    # 판독 상호와 폴더명 기준 업체명 양쪽으로 기존 거래처를 찾는다.
    # (예: 판독 '미가건축인테리어' ↔ ERP '미가건축')
    prev = by_name.get(norm(name)) or by_name.get(norm(FMAP[folder]['name']))
    try:
        if prev:
            # 기존 값이 있으면 유지. 빈칸만 채운다.
            patch = {k: v for k, v in payload.items()
                     if v and not str(prev.get(k) or '').strip()}
            # 덮어쓰지는 않지만, 서류와 기존 값이 다르면 알려야 한다.
            for k in ('business_number', 'address', 'bank_info', 'representative'):
                old, new = str(prev.get(k) or '').strip(), str(payload.get(k) or '').strip()
                if old and new and squash(old) != squash(new):
                    log['conflicts'].append({'name': name, 'field': k,
                                             'erp_기존': old, '서류_판독': new})
            vid = prev['id']
            if patch and APPLY:
                api('PATCH', f'/rest/v1/vendors?id=eq.{vid}', patch, {'Prefer': 'return=minimal'})
            log['updated'].append({'id': vid, 'name': name, 'filled': sorted(patch)})
            action = f'병합(빈칸 {len(patch)}개 채움)'
        else:
            if APPLY:
                created = api('POST', '/rest/v1/vendors', payload, {'Prefer': 'return=representation'})
                vid = created[0]['id']
            else:
                vid = None
            log['inserted'].append({'id': vid, 'name': name})
            action = '신규'

        # --- 서류 업로드 ---
        urls, up = {}, []
        seq = 0
        done_src = {}  # 합본 파일(사업자등록증+통장사본)은 한 번만 올리고 URL을 나눠 쓴다
        for tag, field in SLOTS:
            if prev and str(prev.get(field) or '').strip():
                continue  # 기존 첨부가 있으면 건드리지 않는다
            f, rejected = pick_for_slot(folder, name, tag)
            for fname, whose in rejected:
                log['rejected_files'].append({'vendor': name, 'slot': field,
                                              'file': fname, '다른업체명': whose})
            if not f:
                continue
            if f['abs'] in done_src:
                urls[field] = done_src[f['abs']]
                up.append(f"{field}←{f['name']}(합본)")
                continue
            seq += 1
            if not vid:
                up.append(f"{field}←{f['name']}")
                continue
            key = safe_storage_path(f"vendors/{vid}/{seq}_{f['name']}")
            blob = Path(f['abs']).read_bytes()
            ctype = mimetypes.guess_type(f['name'])[0] or 'application/octet-stream'
            if APPLY:
                api('POST', f'/storage/v1/object/documents/{key}', raw=blob, ctype=ctype,
                    extra={'x-upsert': 'true'})
            public = f'{URL}/storage/v1/object/public/documents/{key}'
            urls[field] = public
            done_src[f['abs']] = public
            up.append(f"{field}←{f['name']}")
            log['uploaded'].append({'vendor': name, 'field': field, 'path': key,
                                    'source': f['abs'], 'bytes': f['size']})

        if urls and APPLY:
            api('PATCH', f'/rest/v1/vendors?id=eq.{vid}', urls, {'Prefer': 'return=minimal'})

        print(f'  {action:20s} {name:24s} 서류 {len(up)}건 {" ".join(up)}')
    except Exception as e:
        log['errors'].append({'name': name, 'folder': folder, 'error': str(e)})
        print(f'  !! 실패 {name}: {e}')

print(f"\n신규 {len(log['inserted'])} / 병합 {len(log['updated'])} / 서류 {len(log['uploaded'])} / 실패 {len(log['errors'])}")

if log['rejected_files']:
    print(f"\n!! 다른 업체 이름이 박혀 있어 붙이지 않은 파일 {len(log['rejected_files'])}건")
    for x in log['rejected_files']:
        print(f"   {x['vendor']} / {x['slot']} ← {x['file']}  (파일명에 '{x['다른업체명']}')")

if log['conflicts']:
    print(f"\n!! 기존 ERP 값과 서류가 다른 항목 {len(log['conflicts'])}건 (기존 값 유지함)")
    for c in log['conflicts']:
        print(f"   {c['name']} · {c['field']}\n      ERP  : {c['erp_기존']}\n      서류 : {c['서류_판독']}")
if APPLY:
    p = WORK / f"migration-log-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
    p.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'롤백용 로그: {p}')
else:
    print('\n미리보기였습니다. 실제로 넣으려면 --apply 를 붙여 다시 실행하세요.')
